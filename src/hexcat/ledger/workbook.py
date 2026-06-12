"""Write the 5-sheet Excel ledger workbook (openpyxl) — Stage 1 output.

Sheet names and German headers match the operator's Cisco_BUILD_LEDGER.xlsx verbatim:
  Fortschritt | Neue Artikel | Quellen-Tracker | PN-Korrekturen (Feed-ID) | Familien-Audit
Files-only, no database. Written to output/ (never _scratch/).
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

from .engine import SourceResult

SHEET_NEUE = "Neue Artikel"
SHEET_QUELLEN = "Quellen-Tracker"
SHEET_KORR = "PN-Korrekturen (Feed-ID)"
SHEET_AUDIT = "Familien-Audit"
SHEET_FORTSCHRITT = "Fortschritt"

HEADERS_NEUE = [
    "Artikelnummer (Part Number)", "Hauptkategorie", "Unterkategorie",
    "Quelle (Cisco Datasheet)", "Quell-URL", "Verifiziert am", "Notiz",
]
HEADERS_QUELLEN = ["Gruppe", "Cisco Datasheet", "URL", "Status", "Notiz"]
HEADERS_KORR = [
    "In Daten vorhanden", "Korrekte Cisco-PN", "Tab (Daten)",
    "Problem", "bestätigt via Datasheet",
]
HEADERS_AUDIT = ["Bereich", "Familie", "Anzahl PNs", "Status", "Befund / Hinweis"]


def _bold_header(ws, ncols: int, row: int = 1) -> None:
    from openpyxl.styles import Font

    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).font = Font(bold=True)


def build_workbook(results: list[SourceResult], *, run_date: str | None = None):
    import openpyxl
    from openpyxl.styles import Font

    run_date = run_date or date.today().isoformat()
    wb = openpyxl.Workbook()

    # --- Fortschritt (first sheet) -------------------------------------------
    ws = wb.active
    ws.title = SHEET_FORTSCHRITT
    total_pns = sum(len(r.rows) for r in results)
    total_corr = sum(len(r.corrections) for r in results)
    total_flag = sum(len(r.flagged) for r in results)
    new_known = any(r.new_count is not None for r in results)
    total_new = sum((r.new_count or 0) for r in results) if new_known else None
    ws.append(["Cisco Katalog-Aufbau – Fortschritt (Stage 1)"])
    ws.append([])
    ws.append(["Quellen verarbeitet", len(results)])
    ws.append(["PNs gemined (gesamt)", total_pns])
    ws.append(["PN-Korrekturen angewendet", total_corr])
    ws.append(["PNs geflaggt (unbestätigt)", total_flag])
    ws.append(["Neu (vs. Live-Liste)", total_new if total_new is not None
               else "n/a (keine Live-Liste)"])
    ws.append(["Lauf-Datum", run_date])
    ws["A1"].font = Font(bold=True, size=12)

    # --- Neue Artikel --------------------------------------------------------
    ws = wb.create_sheet(SHEET_NEUE)
    ws.append(HEADERS_NEUE)
    _bold_header(ws, len(HEADERS_NEUE))
    for res in results:
        for r in res.rows:
            ws.append([
                r.pn, r.hauptkategorie, r.unterkategorie,
                r.quelle, r.quell_url, r.verifiziert_am, r.notiz,
            ])

    # --- Quellen-Tracker -----------------------------------------------------
    ws = wb.create_sheet(SHEET_QUELLEN)
    ws.append(HEADERS_QUELLEN)
    _bold_header(ws, len(HEADERS_QUELLEN))
    for res in results:
        notiz = f"Tier: {res.tier}; {len(res.rows)} PNs klassifiziert"
        if res.flagged:
            notiz += f"; {len(res.flagged)} geflaggt"
        if res.new_count is None:
            notiz += "; ohne Live-Liste (kein Neu/Bereits-Tag)"
        ws.append([res.source.gruppe, res.source.datasheet, res.source.url,
                   res.status, notiz])

    # --- PN-Korrekturen (Feed-ID) -------------------------------------------
    ws = wb.create_sheet(SHEET_KORR)
    ws.append(HEADERS_KORR)
    _bold_header(ws, len(HEADERS_KORR))
    for res in results:
        for c in res.corrections:
            ws.append([c.raw, c.canonical, c.tab, c.problem, c.confirmed_via])

    # --- Familien-Audit ------------------------------------------------------
    ws = wb.create_sheet(SHEET_AUDIT)
    ws.append([f"Familien-Coverage-Audit (Stage 1) – {run_date}"])
    ws.append([
        f"Diese Charge: {total_pns} eindeutige PNs aus {len(results)} Datasheet(s). "
        "Keine echten Duplikate (Dedup gegen sich selbst)."
    ])
    ws.append([])
    ws.append(HEADERS_AUDIT)
    _bold_header(ws, len(HEADERS_AUDIT), row=4)
    coverage: dict[str, int] = {}
    for res in results:
        for uk, n in res.coverage().items():
            coverage[uk] = coverage.get(uk, 0) + n
    for uk in sorted(coverage):
        ws.append(["Transceivers", uk, coverage[uk], "OK", ""])
    ws["A1"].font = Font(bold=True, size=12)

    return wb


def write_workbook(results: list[SourceResult], out_path: str | Path,
                   *, run_date: str | None = None) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook(results, run_date=run_date)
    wb.save(out)
    return out
