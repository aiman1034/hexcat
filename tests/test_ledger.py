"""Phase 3 Stage-1 ledger engine tests — deterministic, no network.

Covers: ledger spec + locked-22 drift guard, HTML mining (ordering-table targeting +
dedup + token filter), §6 PN normalization acceptance, fetch tier fallback (mocked),
classification buckets, workbook structure, and the §7 skeleton export.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from hexcat.ledger import (
    export_skeleton,
    fetch_datasheet,
    load_ledger_spec,
    normalize_pn,
    run_source,
    verify_ledger_spec,
    write_workbook,
)
from hexcat.ledger import fetch as fetch_mod
from hexcat.ledger.engine import Source
from hexcat.ledger.mine import MinedPN, mine_html
from hexcat.models import INTAKE_COLUMNS

SPEC = load_ledger_spec()

# A tiny datasheet stand-in: a decoy spec table + the ordering table (Description |
# Product Number), with a duplicate row and a junk cell to exercise dedup + token filter.
FIXTURE_HTML = """
<html><body>
<table>
  <tr><th>Parameter</th><th>Symbol</th></tr>
  <tr><td>Wavelength</td><td>850 nm</td></tr>
</table>
<table>
  <tr><th>Description</th><th>Product Number</th></tr>
  <tr><td>10GBASE-SR SFP+ Module for MMF</td><td>SFP-10G-SR</td></tr>
  <tr><td>10GBASE-LR SFP+ Module for SMF</td><td>SFP-10G-LR</td></tr>
  <tr><td>dup row (same PN)</td><td>SFP-10G-SR</td></tr>
  <tr><td>10GBASE-CU SFP+ Cable 3 Meter, passive</td><td>SFP-H10GB-CU3M</td></tr>
  <tr><td>10GBASE-AOC SFP+ Cable 3 Meter</td><td>SFP-10G-AOC3M</td></tr>
  <tr><td>not a part number</td><td>See note 1</td></tr>
</table>
</body></html>
"""


# --- spec + locked-22 drift guard ------------------------------------------------
def test_verify_ledger_spec_passes():
    spec = verify_ledger_spec()
    assert spec.brand == "Cisco"
    assert spec.hauptkategorie == "Transceivers"


def test_classify_buckets():
    assert SPEC.classify_pn("SFP-10G-SR") == "SFP+"
    assert SPEC.classify_pn("SFP-10G-T-X") == "SFP+"          # 10GBASE-T module, NOT a cable
    assert SPEC.classify_pn("SFP-H10GB-CU3M") == "DAC Kabel"
    assert SPEC.classify_pn("SFP-H10GB-ACU7M") == "DAC Kabel"  # active copper -> DAC
    assert SPEC.classify_pn("SFP-10G-AOC3M") == "AOC Kabel"


def test_classify_multi_family():
    """The full Cisco line spans SFP..OSFP — not just the 10G SFP+ pilot. Each form factor
    must land in its own locked-22 bucket (regression for the 35-SKU single-family undercount)."""
    cases = {
        "GLC-T": "SFP", "SFP-1G-SX": "SFP", "SFP-OC3-SR": "SFP",      # 1G / SONET SFP
        "SFP-10G-LR": "SFP+", "SFP-25G-SR-S": "SFP28", "SFP-10/25G-LR-S": "SFP28",
        "SFP-50G-SR-S": "SFP56", "QSFP-40G-SR4": "QSFP+", "WSP-Q40GLR4L": "QSFP+",
        "QSFP-100G-LR4-S": "QSFP28", "QSFP-200G-FR4-S": "QSFP56", "QSFP-400G-DR4": "QSFP112",
        "QDD-400G-DR4-S": "QSFP-DD", "QDD-2X400G-FR4": "QSFP-DD800", "QDD-8X100G-FR": "QSFP-DD800",
        "OSFP-800G-DR8": "OSFP", "CPAK-100G-SR10": "CPAK", "CXP-100G-SR10": "CXP",
        "CFP-100G-LR4": "CFP", "CFP2-100G-ER4": "CFP2", "X2-10GB-SR": "X2",
        "XFP-10G-MM-SR": "XFP", "DWDM-XFP-C": "XFP",
        "QSFP-100G-CU3M": "DAC Kabel", "QSFP-H40G-CU5M": "DAC Kabel",  # copper -> DAC by -CU rule
        "QSFP-4X10G-AC10M": "DAC Kabel", "QSFP-100G-AOC10M": "AOC Kabel",
    }
    from hexcat.config import load_taxonomy
    locked22 = set(load_taxonomy().subcategories)
    for pn, expected in cases.items():
        assert SPEC.classify_pn(pn) == expected, f"{pn} -> {SPEC.classify_pn(pn)} != {expected}"
        # every emitted bucket must be a real locked-22 token (identity map)
        assert SPEC.to_locked22(expected) in locked22


def test_exclude_flags_non_transceivers():
    """Licenses / RTUs / converter adapters are flagged, not emitted (1000% rule)."""
    assert SPEC.is_excluded("CVR-QSFP-SFP10G", "QSFP 40G to SFP+ 10G Adapter Module")
    assert SPEC.is_excluded("R-PON-10G-RTU", "XGS-PON RTU License")
    assert SPEC.is_excluded("CFP2-LIC-UPG-200G", "License for WDM Digital CFP2 enabling 200G")
    # a real transceiver whose description merely mentions an adapter is NOT excluded
    assert not SPEC.is_excluded("CPAK-100G-LR4", "Cisco CPAK Module ... includes LC-SC adapter")
    assert not SPEC.is_excluded("SFP-10G-SR", "10GBASE-SR SFP+ Module for MMF")


def test_locked22_mapping():
    assert SPEC.to_locked22("SFP+") == "SFP+"
    assert SPEC.to_locked22("DAC Kabel") == "DAC Kabel"
    assert SPEC.to_locked22("AOC Kabel") == "AOC Kabel"
    assert SPEC.to_locked22("XFP") == "XFP"  # identity for unlisted


# --- mining ----------------------------------------------------------------------
def test_mine_html_targets_ordering_table_and_dedups():
    mined = mine_html(FIXTURE_HTML, SPEC)
    pns = [m.pn for m in mined]
    assert pns == ["SFP-10G-SR", "SFP-10G-LR", "SFP-H10GB-CU3M", "SFP-10G-AOC3M"]
    assert "See note 1" not in pns          # token filter drops prose
    assert pns.count("SFP-10G-SR") == 1      # dedup against itself
    assert mined[0].description.startswith("10GBASE-SR")


# --- §6 normalization acceptance (grounded in the operator's own PN-Korrekturen) --
CANONICAL = {"SFP-10G-SR", "SFP-10G-T-X", "SFP-10G-LR"}


def test_norm_feed_id_suffix_stripped():
    r = normalize_pn("SFP-10G-SR-P-4683", CANONICAL, SPEC)
    assert r.canonical == "SFP-10G-SR"
    assert r.confirmed and r.is_correction
    assert r.problem == "Feed-ID-Anhang -P-####"


def test_norm_spelling_missing_hyphen():
    r = normalize_pn("SFP-10G-TX", CANONICAL, SPEC)
    assert r.canonical == "SFP-10G-T-X"
    assert r.confirmed and r.is_correction
    assert r.problem == "Schreibweise (fehlender Bindestrich)"


def test_norm_already_clean_no_correction():
    r = normalize_pn("SFP-10G-SR", CANONICAL, SPEC)
    assert r.canonical == "SFP-10G-SR"
    assert not r.is_correction and r.confirmed


def test_norm_unconfirmed_is_flagged_not_forced():
    # Stripping yields a form NOT in the datasheet set -> changed but unconfirmed.
    r = normalize_pn("SFP-10G-BOGUS-P-99", CANONICAL, SPEC)
    assert r.is_correction          # the suffix was stripped
    assert not r.confirmed          # but the result isn't datasheet-confirmed -> flag


# --- fetch tier fallback (mocked; no network) ------------------------------------
def test_fetch_tier1_success_writes_cache(tmp_path, monkeypatch):
    monkeypatch.setattr(fetch_mod, "CACHE_DIR", tmp_path / "cache")
    monkeypatch.setattr(fetch_mod, "MANUAL_DIR", tmp_path / "manual")

    def fake_get(url):
        return 200, b"<html>ok</html>", "text/html;charset=utf-8"

    res = fetch_datasheet("https://x/data_sheet_c78-455693.html", http_get=fake_get)
    assert res.tier == "tier1-httpx" and res.content_type == "html"
    assert res.path.exists() and res.read_text() == "<html>ok</html>"

    # Second call hits the cache (no getter needed).
    res2 = fetch_datasheet("https://x/data_sheet_c78-455693.html",
                           http_get=lambda u: (_ for _ in ()).throw(AssertionError("should not fetch")))
    assert res2.tier == "cache"


def test_fetch_falls_back_to_manual_dropin(tmp_path, monkeypatch):
    monkeypatch.setattr(fetch_mod, "CACHE_DIR", tmp_path / "cache")
    manual_dir = tmp_path / "manual"
    manual_dir.mkdir()
    monkeypatch.setattr(fetch_mod, "MANUAL_DIR", manual_dir)
    (manual_dir / "c78-999999.html").write_text("<html>manual</html>", encoding="utf-8")

    def boom(url):
        raise RuntimeError("blocked 403")

    res = fetch_datasheet("https://x/data_sheet_c78-999999.html", http_get=boom)
    assert res.tier == "tier3-manual"
    assert res.read_text() == "<html>manual</html>"


def test_fetch_raises_when_all_tiers_fail(tmp_path, monkeypatch):
    monkeypatch.setattr(fetch_mod, "CACHE_DIR", tmp_path / "cache")
    monkeypatch.setattr(fetch_mod, "MANUAL_DIR", tmp_path / "manual")
    with pytest.raises(fetch_mod.FetchError):
        fetch_datasheet("https://x/data_sheet_c78-000000.html",
                        http_get=lambda u: (500, b"", "text/html"))


# --- engine + workbook + export end-to-end (on the fixture) ----------------------
def _fixture_result():
    src = Source(gruppe="10G", datasheet="Cisco 10GBASE SFP+ Modules",
                 url="https://x/data_sheet_c78-455693.html")
    fetched = type("F", (), {"tier": "fixture", "content_type": "html",
                             "read_text": lambda self: FIXTURE_HTML})()
    return run_source(src, SPEC, verified_date="2026-06-12", fetched=fetched)


def test_run_source_classifies_and_counts():
    res = _fixture_result()
    assert res.mined_count == 4
    assert res.coverage() == {"SFP+": 2, "DAC Kabel": 1, "AOC Kabel": 1}
    assert res.new_count is None  # no live list


def test_run_source_with_live_list_tags_and_corrects():
    src = Source(gruppe="10G", datasheet="Cisco 10GBASE SFP+ Modules",
                 url="https://x/data_sheet_c78-455693.html")
    fetched = type("F", (), {"tier": "fixture", "content_type": "html",
                             "read_text": lambda self: FIXTURE_HTML})()
    # live feed: one messy (feed-id) existing PN + one clean existing PN.
    res = run_source(src, SPEC, verified_date="2026-06-12",
                     live_pns=["SFP-10G-SR-P-4683", "SFP-10G-LR"], fetched=fetched)
    assert res.new_count == 2  # CU3M + AOC3M are new; SR + LR already live
    corr = {c.raw: c for c in res.corrections}
    assert "SFP-10G-SR-P-4683" in corr
    assert corr["SFP-10G-SR-P-4683"].canonical == "SFP-10G-SR"
    assert corr["SFP-10G-SR-P-4683"].tab == "SFP+"
    assert "C78-455693" in corr["SFP-10G-SR-P-4683"].confirmed_via


def test_workbook_structure(tmp_path):
    import openpyxl

    res = _fixture_result()
    out = write_workbook([res], tmp_path / "ledger.xlsx", run_date="2026-06-12",
                         brand=SPEC.brand)
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == [
        "Fortschritt", "Neue Artikel", "Quellen-Tracker",
        "PN-Korrekturen (Feed-ID)", "Familien-Audit",
    ]
    neue = wb["Neue Artikel"]
    assert [c.value for c in neue[1]] == [
        "Artikelnummer (Part Number)", "Hauptkategorie", "Unterkategorie",
        "Quelle (Cisco Datasheet)", "Quell-URL", "Verifiziert am", "Notiz",
    ]
    assert neue.max_row == 1 + 4  # header + 4 mined PNs


def test_export_skeleton_maps_to_locked22(tmp_path):
    import csv

    res = _fixture_result()
    out = export_skeleton([res], tmp_path / "skeleton.csv", SPEC)
    with out.open(encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert list(rows[0].keys()) == list(INTAKE_COLUMNS)
    by_pn = {r["Artikelnummer"]: r for r in rows}
    assert by_pn["SFP-10G-SR"]["KategorieEbene3"] == "SFP+"        # mapped from "SFP+ (10G)"
    assert by_pn["SFP-H10GB-CU3M"]["KategorieEbene3"] == "DAC Kabel"
    assert by_pn["SFP-10G-SR"]["Vendor"] == "Cisco"
    assert by_pn["SFP-10G-SR"]["Kurzbeschreibung"] == ""           # specs/prose stay blank
    assert by_pn["SFP-10G-SR"]["SourceURLs"].endswith("c78-455693.html")


# --- PDF mining (Phase 4): committed fixture, no network -------------------------
FIXTURES = Path(__file__).resolve().parent / "fixtures"
SAMPLE_PDF = FIXTURES / "sample_datasheet.pdf"


def _pdf_spec(pdf_cfg: dict):
    """A minimal but valid LedgerSpec carrying just a mine.pdf block for the fixture."""
    from hexcat.ledger.spec import LedgerSpec

    return LedgerSpec.model_validate({
        "brand": "Fixture",
        "hauptkategorie": "Transceivers",
        "mine": {"pn_header_patterns": ["sku"], "pn_token": r"^.+$", "pdf": pdf_cfg},
        "normalize": {"feed_id_suffix": r"-P-[0-9]+$", "problem_feed_id": "x",
                      "spelling_fixes": {}, "problem_spelling": ""},
        "classify": {"rules": [], "default": "SFP"},
        "locked22_map": {},
    })


def test_mine_pdf_token_scopes_and_scans():
    """Token mode: scope to the 'Ordering Information' page and keep SKU-shaped tokens.
    The section-only page 2 carries no scope heading, so its SKUs must not leak in."""
    from hexcat.ledger.mine import mine_pdf

    spec = _pdf_spec({
        "mode": "token",
        "scope_heading": "Ordering Information",
        "sku_token": r"(?:FN|FG|FR|SP)-[A-Z0-9]+(?:[-+][A-Z0-9]+)+",
    })
    pns = [m.pn for m in mine_pdf(SAMPLE_PDF.read_bytes(), spec)]
    assert pns == ["FN-TRAN-SFP+LR", "FN-TRAN-QSFP28-SR4", "FG-CABLE-SR10-SFP"]


def test_mine_pdf_section_heading_guard_adjacency_and_module_exclusion():
    """Section mode in one shot proves the four hard parts:
      * the bold double-rendered heading switches the chapter (form factor = SFP+),
      * the non-doubled contents line 'SFP+ Modules listed on page 2' does NOT,
      * an adjacent '<noun> (SKU)' optic callout is kept and tagged from the heading,
      * a 'Module (SKU)' callout is dropped (Module = switch line card, not an optic)."""
    from hexcat.ledger.mine import mine_pdf

    spec = _pdf_spec({
        "mode": "section",
        "sku_token": r"(?:[A-Z][0-9][A-Z][0-9]{2}[A-Z]|J[A-Z][0-9]{3}[A-Z])",
        "context_noun": r"(?:[Tt]ransceiver|AOC|DAC)",
        "collapse_bold_doubling": True,
        "chapters": {"SFP+ Modules": "SFP+"},
    })
    mined = mine_pdf(SAMPLE_PDF.read_bytes(), spec)
    assert [(m.pn, m.unterkategorie) for m in mined] == [("R0Z21A", "SFP+")]
    assert "JL999A" not in {m.pn for m in mined}  # switch-line-card 'Module (SKU)' excluded


def test_mine_pdf_section_emits_locked22_through_engine():
    """The section-mode chapter tag must flow through run_source as the Unterkategorie
    (engine prefers the mine-time form factor over from-PN classification)."""
    spec = _pdf_spec({
        "mode": "section",
        "sku_token": r"(?:[A-Z][0-9][A-Z][0-9]{2}[A-Z]|J[A-Z][0-9]{3}[A-Z])",
        "context_noun": r"(?:[Tt]ransceiver|AOC|DAC)",
        "collapse_bold_doubling": True,
        "chapters": {"SFP+ Modules": "SFP+"},
    })
    src = Source(gruppe="Fixture", datasheet="Fixture Guide", url="https://x/sample.pdf")
    fetched = type("F", (), {"tier": "fixture", "content_type": "pdf",
                             "read_bytes": lambda self: SAMPLE_PDF.read_bytes()})()
    res = run_source(src, spec, verified_date="2026-06-12", fetched=fetched)
    assert res.mined_count == 1
    assert res.rows[0].unterkategorie == "SFP+"


def test_hpe_section_spec_verifies_against_locked22():
    """The shipped HPE/Aruba section spec must pass the locked-22 honesty check — including
    its chapter form-factor tags, which are emitted directly (bypassing classify)."""
    from hexcat.config import load_taxonomy
    from hexcat.ledger.spec import LEDGER_CONFIG_DIR, verify_ledger_spec

    spec = verify_ledger_spec(str(LEDGER_CONFIG_DIR / "hpe_aruba_transceivers.yaml"))
    assert spec.mine.pdf is not None and spec.mine.pdf.mode == "section"
    locked22 = set(load_taxonomy().subcategories)
    assert set(spec.mine.pdf.chapters.values()) <= locked22


def test_verify_rejects_section_chapter_outside_locked22(tmp_path):
    """A section-mode chapter tag that is not a locked-22 token must fail loudly — otherwise
    a section PDF could emit an Unterkategorie invalid for the Stage-3 hand-off."""
    from hexcat.ledger.spec import LedgerSpecError, verify_ledger_spec

    bad = tmp_path / "bad_section.yaml"
    bad.write_text(
        "brand: X\nhauptkategorie: Transceivers\n"
        "mine:\n  pn_header_patterns: ['sku']\n  pn_token: '^.+$'\n"
        "  pdf:\n    mode: section\n    sku_token: 'J[A-Z][0-9]{3}[A-Z]'\n"
        "    context_noun: 'Transceiver'\n    chapters: {'Bogus Modules': 'NOT-A-FORMFACTOR'}\n"
        "normalize:\n  feed_id_suffix: '-P-[0-9]+$'\n  problem_feed_id: x\n"
        "  spelling_fixes: {}\n  problem_spelling: ''\n"
        "classify:\n  rules: []\n  default: 'SFP'\nlocked22_map: {}\n",
        encoding="utf-8",
    )
    with pytest.raises(LedgerSpecError):
        verify_ledger_spec(str(bad))


# --- optional integration: real cached datasheet, if present ---------------------
def test_real_datasheet_yields_35_if_cached():
    cached = Path(__file__).resolve().parents[1] / "datasheets" / "cache" / "c78-455693.html"
    if not cached.exists():
        pytest.skip("cached datasheet not present (network/cache-dependent)")
    mined = mine_html(cached.read_text(encoding="utf-8"), SPEC)
    pns = {m.pn for m in mined}
    assert len(mined) == 35
    for expected in ["SFP-10G-SR", "SFP-10G-LR", "SFP-10G-ER", "SFP-10G-LRM",
                     "SFP-10G-ZR", "SFP-10G-T-X", "SFP-H10GB-CU3M", "SFP-10G-AOC3M"]:
        assert expected in pns


# =====================================================================================
# Stage-1 SELF-AUDIT regression suite — one permanent fixture per defect class.
# Backed by a COMMITTED column-layout fixture (sample_ordering_columns.pdf) so these run
# in CI without the gitignored real datasheets. Bug classes locked here:
#   #1 description-bleed phantoms      -> column extraction excludes them (V2)
#   #2 trailing-'+' separator stripping-> preserved, base & '+' stay distinct (V3/V4)
#   #3 MPO-vs-DAC misclassification    -> universal V5 from description (V5)
#   #4 pack-of-four context            -> Notiz flag
# plus the verifier's V1-V8 calibration frozen on synthetic tokens.
# =====================================================================================
SAMPLE_COLUMNS_PDF = FIXTURES / "sample_ordering_columns.pdf"


def _columns_spec():
    """LedgerSpec for the column fixture: column-aware token mode + a SFP+ default so cables
    are forced to be classified by the universal V5 description rule, not a PN substring."""
    from hexcat.ledger.spec import LedgerSpec

    return LedgerSpec.model_validate({
        "brand": "Fixture", "hauptkategorie": "Transceivers",
        "mine": {"pn_header_patterns": ["sku"], "pn_token": r"^.+$", "pdf": {
            "mode": "token", "scope_heading": "Ordering Information",
            "sku_token": r"FX-[A-Z0-9]+(?:[-+][A-Z0-9]+)+\+?",
            "sku_column": "SKU", "desc_column": "Description"}},
        "normalize": {"feed_id_suffix": r"-P-[0-9]+$", "problem_feed_id": "x",
                      "spelling_fixes": {}, "problem_spelling": ""},
        "classify": {"rules": [], "default": "SFP+"}, "locked22_map": {},
    })


def test_classify_cable_from_description_universal_rule():
    from hexcat.ledger.spec import classify_cable_from_description as cc

    assert cc("10 GE SFP+ passive direct attach cable, 1m") == "DAC Kabel"
    assert cc("10 GE SFP+ active direct attach cable, 10m") == "DAC Kabel"  # active copper, NOT AOC
    assert cc("400 GE QSFPDD active optical cable, 3m") == "AOC Kabel"
    assert cc("100 GE parallel breakout MPO to 10xLC connectors, OM3 MMF") == "MPO Kabel"
    assert cc("40 GE QSFP+ transceiver module, MPO-12 connector") is None  # connector != breakout
    assert cc("DAC") == "DAC Kabel"   # HPE-style bare-noun callout label
    assert cc("AOC") == "AOC Kabel"
    assert cc("Transceiver") is None
    assert cc("") is None


def test_mine_columns_excludes_description_phantom():
    """#1: a SKU-shaped token bleeding through the Description column must NOT be mined."""
    from hexcat.ledger.mine import mine_pdf

    mined = mine_pdf(SAMPLE_COLUMNS_PDF.read_bytes(), _columns_spec())
    pns = {m.pn for m in mined}
    assert "FX-PHANTOM-XX" not in pns           # the description-prose phantom is excluded
    assert "FX-TRAN-NOTE" in pns                # its real SKU-column neighbour is kept


def test_mine_columns_preserves_trailing_plus_distinct():
    """#2: FX-CABLE-DAC1 and FX-CABLE-DAC1+ are DISTINCT (the DR4/DR4+ class) — no '+' drop,
    no silent collision."""
    from hexcat.ledger.mine import mine_pdf

    pns = {m.pn for m in mine_pdf(SAMPLE_COLUMNS_PDF.read_bytes(), _columns_spec())}
    assert "FX-CABLE-DAC1" in pns and "FX-CABLE-DAC1+" in pns


def test_mine_columns_v5_classifies_cables_from_description():
    """#3: classification comes from the manufacturer description, not the PN substring."""
    from hexcat.ledger.mine import mine_pdf

    spec = _columns_spec()
    uk = {m.pn: spec.resolve_unterkategorie(m.pn, description=m.description,
                                            hint=m.unterkategorie)
          for m in mine_pdf(SAMPLE_COLUMNS_PDF.read_bytes(), spec)}
    assert uk["FX-CABLE-DAC1"] == "DAC Kabel"
    assert uk["FX-CABLE-SR10"] == "MPO Kabel"   # OM3 MPO breakout, despite no 'MPO' in the PN
    assert uk["FX-CABLE-AOC03"] == "AOC Kabel"
    assert uk["FX-TRAN-SFP+LR"] == "SFP+"        # plain transceiver -> form factor


def test_pack_of_four_flagged_in_notiz():
    """#4: a Pack-of-four SKU carries a Viererpack Notiz through the engine."""
    spec = _columns_spec()
    src = Source(gruppe="Fixture", datasheet="Fixture Guide", url="https://x/columns.pdf")
    fetched = type("F", (), {"tier": "fixture", "content_type": "pdf",
                             "read_bytes": lambda self: SAMPLE_COLUMNS_PDF.read_bytes()})()
    res = run_source(src, spec, verified_date="2026-06-12", fetched=fetched)
    notiz = {r.pn: r.notiz for r in res.rows}
    assert "Viererpack" in notiz["FX-TRAN-SX-4PACK"]
    assert notiz["FX-CABLE-DAC1"] == ""          # non-pack rows carry no pack flag


# --- verifier V1-V8 calibration, frozen on synthetic tokens --------------------------
def _emitted(pairs):
    from hexcat.verify.checks import EmittedRow

    return [EmittedRow(pn=p, unterkategorie=u, notiz="") for p, u in pairs]


def _tok(pn, desc="", locus="sku_column"):
    from hexcat.verify.extract import SourceToken

    return SourceToken(pn=pn, description=desc, locus=locus, page=1)


def test_verifier_v2_flags_description_phantom():
    from hexcat.verify import checks as C

    tokens = [_tok("FX-TRAN-LR", "transceiver"),
              _tok("FX-PHANTOM-XX", "see FX-PHANTOM-XX", locus="description")]
    emitted = _emitted([("FX-TRAN-LR", "SFP+"), ("FX-PHANTOM-XX", "SFP+")])
    res = C.v2_provenance(emitted, tokens)
    assert not res.passed and res.offenders == ["FX-PHANTOM-XX"]


def test_verifier_v1_v4_flag_trailing_plus_mangle():
    from hexcat.verify import checks as C

    tokens = [_tok("SP-CABLE-ADASFP+", "active direct attach cable")]
    raw = "ordering SP-CABLE-ADASFP+ 10 GE active direct attach cable"
    emitted = _emitted([("SP-CABLE-ADASFP", "DAC Kabel")])  # '+' stripped
    v1 = C.v1_verbatim(emitted, raw, ["SP-CABLE-ADASFP+"])
    v4 = C.v4_separator_integrity(emitted, tokens)
    assert not v1.passed and "SP-CABLE-ADASFP" in v1.offenders
    assert not v4.passed and "SP-CABLE-ADASFP" in v4.offenders


def test_verifier_v3_flags_silent_collision_and_passes_when_both_kept():
    from hexcat.verify import checks as C
    from hexcat.ledger.spec import load_ledger_spec

    spec = load_ledger_spec()
    tokens = [_tok("FN-TRAN-QSFPDD-DR4", "500m"), _tok("FN-TRAN-QSFPDD-DR4+", "2km")]
    lost = _emitted([("FN-TRAN-QSFPDD-DR4", "QSFP-DD")])          # '+' form silently deduped
    both = _emitted([("FN-TRAN-QSFPDD-DR4", "QSFP-DD"),
                     ("FN-TRAN-QSFPDD-DR4+", "QSFP-DD")])
    assert not C.v3_no_silent_collision(lost, tokens, spec).passed
    assert C.v3_no_silent_collision(both, tokens, spec).passed   # no false positive when kept


def test_verifier_v4_no_false_positive_when_both_forms_emitted():
    from hexcat.verify import checks as C

    tokens = [_tok("FN-TRAN-QSFPDD-DR4", "500m"), _tok("FN-TRAN-QSFPDD-DR4+", "2km")]
    emitted = _emitted([("FN-TRAN-QSFPDD-DR4", "QSFP-DD"),
                        ("FN-TRAN-QSFPDD-DR4+", "QSFP-DD")])
    assert C.v4_separator_integrity(emitted, tokens).passed  # DR4 is legit, not a mangle


def test_verifier_v5_flags_mpo_misclassified_as_dac():
    from hexcat.verify import checks as C

    tokens = [_tok("FG-CABLE-SR10", "100 GE parallel breakout MPO to 10xLC connectors, OM3 MMF")]
    emitted = _emitted([("FG-CABLE-SR10", "DAC Kabel")])   # wrong: it's an MPO breakout
    res = C.v5_classification(emitted, tokens)
    assert not res.passed and res.offenders == ["FG-CABLE-SR10"]


def test_verifier_all_green_on_clean_set():
    """True negative: a faithfully-mined set passes every check."""
    from hexcat.verify import checks as C
    from hexcat.ledger.spec import load_ledger_spec

    spec = load_ledger_spec()
    tokens = [_tok("FN-TRAN-SFP+LR", "10 GE SFP+ transceiver module"),
              _tok("FN-CABLE-SFP+1", "10 GE SFP+ passive direct attach cable")]
    raw = "FN-TRAN-SFP+LR 10 GE SFP+ transceiver module FN-CABLE-SFP+1 passive direct attach cable"
    emitted = _emitted([("FN-TRAN-SFP+LR", "SFP+"), ("FN-CABLE-SFP+1", "DAC Kabel")])
    skus = ["FN-TRAN-SFP+LR", "FN-CABLE-SFP+1"]
    v2 = C.v2_provenance(emitted, tokens)
    v4 = C.v4_separator_integrity(emitted, tokens)
    checks = [
        C.v1_verbatim(emitted, raw, skus), v2,
        C.v3_no_silent_collision(emitted, tokens, spec), v4,
        C.v5_classification(emitted, tokens), C.v6_switch_exclusion(emitted, tokens),
        C.v7_completeness(emitted, tokens, v2, v4), C.v8_count_honesty(emitted, tokens),
    ]
    assert all(c.passed for c in checks), [c.name for c in checks if not c.passed]


def test_verify_column_fixture_end_to_end_passes():
    """The committed column fixture, mined + classified through the engine, must verify green."""
    from hexcat.ledger.mine import mine_pdf
    from hexcat.verify.checks import EmittedRow
    from hexcat.verify.verifier import verify_ledger

    spec = _columns_spec()
    data = SAMPLE_COLUMNS_PDF.read_bytes()
    emitted = [
        EmittedRow(pn=m.pn,
                   unterkategorie=spec.resolve_unterkategorie(
                       m.pn, description=m.description, hint=m.unterkategorie),
                   notiz="")
        for m in mine_pdf(data, spec)
    ]
    res = verify_ledger(emitted, spec, brand="Fixture", pdf_bytes=data)
    assert res.passed, [c.name for c in res.checks if not c.passed]
    assert res.authoritative_count == 7


# --- V9 catalog coverage (whole-brand, merged ledger) --------------------------------
def test_v9_passes_when_every_expected_family_present():
    from hexcat.verify import checks as C

    expected = ["SFP", "SFP+", "QSFP28", "DAC Kabel"]
    emitted = _emitted([("GLC-T", "SFP"), ("SFP-10G-SR", "SFP+"),
                        ("QSFP-100G-LR4-S", "QSFP28"), ("QSFP-100G-CU3M", "DAC Kabel")])
    res = C.v9_catalog_coverage(emitted, expected)
    assert res.passed and not res.offenders
    assert res.details["skus_per_expected_family"]["SFP+"] == 1


def test_v9_flags_and_names_missing_family():
    """A family with zero emitted SKUs marks the catalog KNOWN-INCOMPLETE and is named."""
    from hexcat.verify import checks as C

    expected = ["SFP", "SFP+", "QSFP28", "OSFP"]
    emitted = _emitted([("GLC-T", "SFP"), ("SFP-10G-SR", "SFP+"),
                        ("QSFP-100G-LR4-S", "QSFP28")])           # no OSFP
    res = C.v9_catalog_coverage(emitted, expected)
    assert not res.passed
    assert res.offenders == ["OSFP"]
    assert "OSFP" in res.summary


def test_v9_uncalibrated_when_spec_has_no_coverage(tmp_path):
    """A spec with no coverage.expected_families cannot certify DONE-VERIFIED -> V9 FAILs."""
    from hexcat.verify.verifier import verify_catalog_coverage

    spec = load_ledger_spec()
    spec_no_cov = spec.model_copy(update={"coverage": None})
    emitted = _emitted([("SFP-10G-SR", "SFP+")])
    res = verify_catalog_coverage(emitted, spec_no_cov, brand="X")
    assert not res.passed
    assert "UNCALIBRATED" in res.checks[0].summary


def test_cisco_spec_coverage_19_families_all_locked22_and_reachable():
    """The Cisco V9 contract is the 19 known families; verify_ledger_spec already guards that
    each is a locked-22 token AND reachable by a classify rule, so a clean load proves both."""
    spec = verify_ledger_spec()
    assert spec.coverage is not None
    assert len(spec.coverage.expected_families) == 19
    from hexcat.config import load_taxonomy
    locked22 = set(load_taxonomy().subcategories)
    assert all(f in locked22 for f in spec.coverage.expected_families)
